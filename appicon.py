#!/usr/bin/env python

import glob, os
import requests
from zipfile import ZipFile
from PIL import Image
from StringIO import StringIO
from slugify import slugify
from openpyxl import Workbook

debug=0

def write_excell(apps):
    """
    """
    dest_filename = "orange_iphone_description.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "iOS Orange Apps"

    row = 1

    for title, version, bundle_id, description, icon_url in apps:
        ws.cell('A%s' % row).value = title
        ws.cell('B%s' % row).value = version
        ws.cell('C%s' % row).value = bundle_id
        ws.cell('D%s' % row).value = description
        ws.cell('E%s' % row).value = icon_url

        row += 1

    wb.save(filename=dest_filename)


def download_ios_app(app_id, mask, mask512):
    """
    Download iOS app metadata.
    """
    if debug:
        print("======================================================= [{}] ({}) =======================================================".format(app_id[0], app_id[1]))
    url = "http://itunes.apple.com/{}/lookup?id={}".format(app_id[1], app_id[0])
    r = requests.get(url)
    if r.status_code != 200:
        print('Error downloading {} {}'.format(url, r.status_code))
        return

    results = r.json()

    meta = results["results"][0]
    icon_url = meta["artworkUrl512"]
    title = meta["trackCensoredName"]
    version = meta["version"]
    slug = slugify(title)
    description = meta["description"]
    bundle_id = meta["bundleId"]
    
    if debug:
        print("=> download application AdamID={} name={} country={}".format(app_id[0], slug, app_id[1]))

    screenshotiPhone_url = {}
    for nb_screenshots in range(5):
        try:
            screenshotiPhone_url[nb_screenshots] = meta["screenshotUrls"][nb_screenshots]
            if debug:
                print('Screenshot iPhone #{} = {}'.format(nb_screenshots, screenshotiPhone_url[nb_screenshots]))
            download_ios_screenshot(screenshotiPhone_url[nb_screenshots], slug, nb_screenshots)
        except IndexError:
            if debug:
                print('No iPhone screenshot #{}'.format(nb_screenshots))
            pass
    screenshotiPad_url = {}
    for nb_screenshots in range(5):
        try:
            screenshotiPad_url[nb_screenshots] = meta["ipadScreenshotUrls"][nb_screenshots]
            if debug:
                print('Screenshot iPad #{}   = {}'.format(nb_screenshots, screenshotiPad_url[nb_screenshots]))
            download_ios_iPadscreenshot(screenshotiPad_url[nb_screenshots], slug, nb_screenshots)
        except IndexError:
            if debug:
                print('No iPad screenshot #{}'.format(nb_screenshots))
            pass

    if debug:
        print(' ')

    download_ios_icon(icon_url, slug, mask, mask512)

    return title, version, bundle_id, description, icon_url


def download_ios_icon_mask():
    """
    """
    # Download the mask from Dropbox, this way we don't
    # have to provide mask.png and the script is self contained.
    if debug:
        print("download 1024 icon image mask")

    mask_url = "https://raw.githubusercontent.com/manbolo/appicon/master/mask.png"
    mask_data = requests.get(mask_url)
    mask = Image.open(StringIO(mask_data.content))
    mask = mask.convert('L')
    return mask


def download_ios_icon_mask512():
    """
    """
    # Download the mask from github, this way we don't
    # have to provide mask512.png and the script is self contained.
    if debug:
        print("download 512 icon image mask")

    mask512_url = "https://raw.githubusercontent.com/manbolo/appicon/master/mask512.png"
    mask512_data = requests.get(mask512_url)
    mask512 = Image.open(StringIO(mask512_data.content))
    mask512 = mask512.convert('L')
    return mask512


def download_ios_icon(icon_url, prefix, mask=None, mask512=None):
    """
    """
    # Download icon an apply mask.
    icon_data = requests.get(icon_url)
    icon = Image.open(StringIO(icon_data.content))

    if mask:
        try:
            icon.putalpha(mask)
        except:
            if debug:
                print('Oops app seems to be very old. Using 512 icon mask as a fallback!!!')
            if mask512:
                icon.putalpha(mask512)
            pass

    # Compute and save thumbnails.
    for size in [1024, 512, 240, 216, 120, 114, 60, 57]:
        icon_resized = icon.resize((size, size), Image.ANTIALIAS)
        icon_resized.save("icon_{0}_{1}x{1}.png".format(prefix, size))


def download_ios_screenshot(screenshot_url, title, index):
    """
    """
    data = requests.get(screenshot_url)
    img = Image.open(StringIO(data.content))

    width_src, height_src = img.size

    if width_src > height_src:
        width_dsts = [1136, 568]
    else:
        width_dsts = [568, 284]

    for width_dst in width_dsts:
        height_dst = width_dst * height_src / width_src
        img_resized = img.resize((width_dst, height_dst), Image.ANTIALIAS)
        img_resized.save("screenshot_{0}_{1}_{2}.png".format(title, index, width_dst))


def download_ios_iPadscreenshot(screenshot_url, title, index):
    """
    """
    data = requests.get(screenshot_url)
    img = Image.open(StringIO(data.content))

    width_src, height_src = img.size

    if width_src > height_src:
        width_dsts = [480, 240]
    else:
        width_dsts = [360, 180]

    for width_dst in width_dsts:
        height_dst = width_dst * height_src / width_src
        img_resized = img.resize((width_dst, height_dst), Image.ANTIALIAS)
        img_resized.save("screenshot_{0}_{1}_{2}.png".format(title, index, width_dst))


def clean_files():
    """
    """
    for ext in ["*.png", "*.jpg", "*.jpeg"]:
        [os.remove(f) for f in glob.glob(ext)]


def zip_files(archive, pattern):
    """
    """
    with ZipFile(archive, "w") as myzip:
        for f in glob.glob(pattern):
            myzip.write(f)


def download_iphone_img():
    """
    """
    app_ids = [
        ["337707649", "fr"], # memorylife (Orange Vall?e)
        ["402167427", "fr"], # LibonFreecallsSmartVoicemailandInstantMessaging (Orange Vall?e)
        ["336760681", "gb"], # YourOrange (Orange UK)
        ["385011850", "gb"], # OrangeWednesdaysforiPad (Orange UK)
        ["542188355", "gb"], # U24 (Orange UK)
        ["335804093", "gb"], # OrangeWednesdays (Orange UK)
        ["529079503", "gb"], # OneStopfromOrangeUK (Orange UK)
        ["690528311", "gb"], # OrangeCash (Orange UK)
        ["690042828", "es"], # OrangeTV (Orange Spain)
        ["379866108", "es"], # MiOrange (Orange Spain)
        ["582248757", "es"], # DianaOrange (Orange Spain)
        ["848944375", "es"], # OrangeCash (Orange Spain)
        ["760989535", "es"], # OrangeCloud (Orange Spain)
        ["370553432", "es"], # RincndelVago (Orange Spain)
        ["390896393", "ro"], # OrangeTVGo (Orange Romania)
        ["602100905", "ro"], # OrangeCloud (Orange Romania)
        ["420021996", "ro"], # ContulmeuOrange (Orange Romania)
        ["848309908", "ro"], # FotbalInfodelaOrange (Orange Romania)
        ["460371997", "ro"], # TheGoodiesHunt (Orange Romania)
        ["381671309", "ro"], # Orangefilm (Orange Romania)
        ["525384895", "ro"], # OrangeExplorer (Orange Romania)
        ["585617367", "ro"], # InfoSkidelaOrange (Orange Romania)
        ["338908827", "fr"], # actuRunion (Orange R?union)
        ["379867763", "fr"], # sortiesRunion (Orange R?union)
        ["503153960", "pl"], # bramkaSMSOrange (Orange Polska)
        ["663216090", "pl"], # NawigacjaOrange (Orange Polska)
        ["875223114", "pl"], # wiatOrange (Orange Polska)
        ["588720962", "pl"], # MjOrange (Orange Polska)
        ["523132386", "pl"], # TelewizjaTuiTam (Orange Polska)
        ["630465094", "pl"], # WyszukiwarkaSalonwOrange (Orange Polska)
        ["712153878", "pl"], # FunSpot (Orange Polska)
        ["532536991", "pl"], # SerceiRozumjadnaMistrzostwa (Orange Polska)
        ["692682912", "pl"], # PAYBACKOrange (Orange Polska)
        ["871862096", "pl"], # OrangeWarsawFestival (Orange Polska)
        ["806167601", "pl"], # BiznesPakiet (Orange Polska)
        ["685235929", "pl"], # OrangeKinoLetnie (Orange Polska)
        ["394644603", "md"], # OrangeTVMoldova (Orange Moldova SA)
        ["453708836", "md"], # AllUniverse (Orange Moldova SA)
        ["779576547", "lu"], # OrangeCloudLuxembourg (Orange Luxembourg)
        ["399603431", "lu"], # OrangeWednesdaysLuxembourg (Orange Luxembourg)
        ["551949969", "lu"], # MaLiveboxLuxembourg (Orange Luxembourg)
        ["445602028", "jo"], # OrangeBackup (Orange Jordan)
        ["808824590", "jo"], # MyOrangeJordan (Orange Jordan)
        ["823609494", "it"], # OrangeRadio (Orange Horizons)
        ["758009206", "am"], # MyOrangeArmenia (Orange Armenia)
        ["445573616", "fr"], # maLivebox (Orange)
        ["472332600", "fr"], # OrangeBusinessLounge (Orange)
        ["580161193", "fr"], # MailOrange (Orange)
        ["435920353", "fr"], # Fichesant (Orange)
        ["324306873", "fr"], # Ligue1 (Orange)
        ["766413402", "fr"], # OrangePic (Orange)
        ["564725306", "fr"], # DMexpress (Orange)
        ["368541602", "fr"], # ReadandGoMabibliothque (Orange)
        ["286597170", "fr"], # wifidOrange (Orange)
        ["667366691", "fr"], # MonrseaumobileOrange (Orange)
        ["367722531", "fr"], # Orangeetmoi (Orange)
        ["474730795", "fr"], # LiveboxPhonevotretlphonefixeparinternettoujoursavecvousappelsenwifilamaisonetdepartoutaccsvotremessagerievocaleetaurenvoidappels (Orange)
        ["432563668", "fr"], # OrangeCinday (Orange)
        ["470868364", "fr"], # NumroUnique (Orange)
        ["708390020", "fr"], # Paiementpro (Orange)
        ["469908244", "fr"], # LeClouddOrangevotreespacedestockageenligneprivetconfidentielpoursauvegardersynchroniseretpartagervosphotosvidosetautrescontenus (Orange)
        ["509534836", "fr"], # joynbyOrange (Orange)
        ["762577365", "fr"], # VidolademandedOrange (Orange)
        ["776026237", "fr"], # leBlocdOrange (Orange)
        ["608200581", "fr"], # TVcommandedOrange (Orange)
        ["338986109", "fr"], # 118712annuairerechercheinverseproparticulier (Orange)
        ["417977933", "fr"], # OrangeJobs (Orange)
        ["437328781", "fr"], # Messageriepro (Orange)
        ["438207597", "fr"], # 118712pouriPad (Orange)
        ["738366742", "fr"], # Wisheo (Orange)
        ["711493886", "fr"], # Orangeexpomuses (Orange)
        ["580109998", "fr"], # VideoMeeting (Orange)
        ["523083734", "fr"], # MyOfficemobile (Orange)
        ["527538606", "fr"], # ProgrammeTVdOrange (Orange)
        ["849800654", "fr"], # LaMaisonConnecteInternetsurMobile (Orange)
        ["553791662", "fr"], # MySosh (Orange)
        ["328085822", "fr"], # OrangeMaps (Orange)
        ["308816822", "fr"], # TVdOrange (Orange)
        ["628448039", "fr"], # Parnasse (Orange)
        ["489844135", "fr"], # SuiviIntervention (Orange)
        ["493983471", "fr"], # StarAfricafootball (Orange)
        ["711435079", "fr"], # mesboutiquesorange (Orange)
        ["449495567", "fr"], # StarAfrica (Orange)
        ["522391632", "ug"], # GroupTalk (Orange)
        ["848331928", "fr"], # FamilyPlace (Orange)
        ["447705993", "fr"], # InterventionsTempsRel (Orange)
        ["567776217", "fr"], # Maisonconnecte (Orange)
        ["773467154", "fr"], # OCSGO (OCS)
        ["813677605", "fr"], # GameofThronesS4CompagnonTVofficiel (OCS)
        ["702585899", "fr"], # MobileCloud (Mobistar)
        ["626702792", "be"], # MyMobistar (Mobistar)
        ["538344747", "eg"], # 8000 (mobinil)
        ["322560653", "eg"], # mBorsa (mobinil)
        ["843297786", "fr"], # 1013RseauSignalisationdundommagetouchantlerseautlphoniquefixe (FTRD UK LTD)
        ["482583090", "fr"], # PhotoonTV (FTRD UK LTD)
        ["802225482", "fr"], # TwinPlayer (FTRD UK LTD)
        ["565097923", "fr"], # BodyGuruLite (FTRD UK LTD)
        ["857660204", "fr"], # CloodStream (FTRD UK LTD)
        ["867444732", "fr"], # radio140 (FTRD UK LTD)
        ["864955324", "fr"], # Magntofut (FTRD UK LTD)
        ["880099588", "fr"], # VibesAround (FTRD UK LTD)
        ["596664884", "fr"], # Streamthatsong (FTRD UK LTD)
        ["489789461", "fr"], # WhatsThatTrack (FTRD UK LTD)
        ["526647373", "fr"], # PillTag (FTRD UK LTD)
        ["583336894", "fr"], # Samplista (FTRD UK LTD)
        ["843193879", "fr"], # ShadesOfHue (FTRD UK LTD)
        ["715573079", "fr"], # MixAround (FTRD UK LTD)
        ["449579959", "fr"], # RglageParabole (FTRD UK LTD)
        ["659408449", "fr"], # PartycallLite (FTRD UK LTD)
        ["882667639", "fr"], # DropOnTV (FTRD UK LTD)
               ]

    clean_files()

    mask = download_ios_icon_mask()
    mask512 = download_ios_icon_mask512()

    apps = [download_ios_app(app_id, mask, mask512) for app_id in app_ids]

    write_excell(apps)

    zip_files("orange_iphone_icon.zip", "icon*.png")

    zip_files("orange_iphone_screenshot.zip", "screenshot*.png")

    clean_files()


if __name__ == '__main__':

    download_iphone_img()
